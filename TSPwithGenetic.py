#R is Initial Populations
#P1 is Parent1
#P2 is Parent2
#N is Neighbors of current nodes in Parent1 and Parent2
#C current node
#L1 is location information of current node in Parent1
#L2 is location information of current node in Parent2
#U is Usable nodes
#PL path length of chromose
#SP shortest path in this population
#CH is created Child
#CT is number of created child
#RN is Row number of shortest path
#SC Stop condition
#M is Mutation
import openpyxl
import numpy as np
import random
from timeit import default_timer as timer

instance = int(input("Please Chose which instance do you want to use(29,42 or 76)"))

wb = openpyxl.load_workbook('Project.xlsx')
start_time = timer()
if instance == 29  :                       
    sheet = wb['Sayfa1']   
elif instance == 42 :
    sheet = wb['Sayfa2']
elif instance == 76 :
    sheet = wb['Sayfa3']
else :
    print("Please write (29-42-76")
R = []
CT = 4999
Chromosome = []
yol = 0
minimum = 10000000000
P1 = 0
P2 = 0
CH = []
N = []
L1 = 0
L2 = 0
PL = 0
SP=50000000
SC = 0
next_node = 0
R = np.zeros((10000000,instance))
PRL = []



for i in range(5000) :
    U=list(range(1, (instance +1 )))
    for j in range(instance):
        next_node = random.choice(U)
        U.remove(int(next_node))
        R[i,j] = int(next_node)
        if j > 0 and j < (instance - 1) :
            yol = yol + sheet.cell(row=R[i][j] + 1, column=R[i][j -1] + 1).value


        elif j == (instance - 1) :
            yol = yol + sheet.cell(row=R[i][j] + 1, column=R[i][j - 1] + 1).value
            yol = yol + sheet.cell(row=R[i][j] + 1, column=R[i][0] + 1).value

    PRL.append(yol)
    yol = 0


    
while SC < 5000 :
    for i in range(10) :
        b = random.randint(0,CT)
        a = PRL[b]        
        if a < minimum :
            minimum = a
            P1 = b
            
    minimum = 100000000

    for i in range(10) :
        b = random.randint(0,CT)
        a = PRL[b]        
        if a < minimum :
            minimum = a
            P2 = b
            
    minimum = 100000000   
    
    
    C = R[P1][0]
    U=list(range(1, (instance +1 )))
    while len(CH) <instance :
        CH.append(C)
        U.remove(C)
        L2 = R[P2].tolist().index(C)


        if  L1 == (instance - 1) :
            N.insert(0,(R[P1][L1 - 1]))
           
        elif L1 == 0 :
            N.insert(0,(R[P1][1]))
            N.insert(1,(R[P1][(instance - 1)]))
        else :
            N.insert(0,(R[P1][L1 - 1]))
            N.insert(1,(R[P1][L1 + 1]))
            
        if L2 == (instance - 1) :
            N.insert(2,(R[P2][L2 - 1]))
            N.insert(3,(R[P2][0]))   
        elif L2 == 0 :
            N.insert(2,(R[P2][L2 + 1]))
            N.insert(3,(R[P1][(instance - 1)]))            
        else :
            N.insert(2,(R[P2][L2 - 1]))
            N.insert(3,(R[P2][L2 + 1])) 
            
        for E in CH:
            while E in N:
                N.remove(E)
                
        if N == [] and U != [] :
            maximum_1 = 10000
            for i in range(len(U)) :
                if sheet.cell(row=CH[-1] + 1, column=U[i] + 1).value < maximum_1:
                    maximum_1 = sheet.cell(row=CH[-1] + 1, column=U[i] + 1).value
                    asdf = U[i]
                
            C = asdf
            

            L1 = R[P1].tolist().index(C) 
            L2 = R[P2].tolist().index(C)  
            N.clear()
            
            
        elif N != []  :
            maximum_1 = 10000
            for i in range(len(N)) :
                if maximum_1 > sheet.cell(row=CH[-1] + 1, column=N[i] + 1).value:
                    maximum_1 = sheet.cell(row=CH[-1] + 1, column=N[i] + 1).value
                    asdf = N[i]
    
            
            C = asdf

            

            L1 = R[P1].tolist().index(C) 
            L2 = R[P2].tolist().index(C) 
            N.clear()
    
    
        

    M = random.randint(0,100)
    if M == 1 :
        xyz = random.randint(0,instance-1)
        zyx = random.randint(0,instance-1)
        CH[xyz],CH[zyx] = CH[zyx],CH[xyz]
    CT = CT + 1
    for i in range(instance):
        R[CT][i] = CH[i]

    
    for i in range(instance):
        if i <(instance - 1) :
                PL = PL + sheet.cell(row=CH[i] + 1, column=CH[i+1] + 1).value

        else : 
                PL = PL + sheet.cell(row=CH[i] + 1, column=CH[0] + 1).value
    PRL.append(PL)
    PL = 0
    

        
    CH.clear()
    if SP > PRL[-1] :
            RN = CT
            SP = PRL[-1]
            SC = 0
    else :
            SC = SC + 1
            
    
print("Shortest pat is : ",SP)
print("Route : ", R[RN])
print("Total child produced : ",CT-5000)    
finish_time = timer() - start_time
print("Run time : ",finish_time)
